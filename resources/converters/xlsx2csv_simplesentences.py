import csv
from openpyxl import load_workbook
import os
import sys

def xlsx_to_csv(input_file, output_file):
    # Load the Excel workbook
    workbook = load_workbook(input_file)
    # Select the first worksheet
    sheet = workbook.active

    # Open output CSV file
    with open(output_file, 'w', newline='') as csv_file:
        # Create CSV writer
        csv_writer = csv.writer(csv_file)
        # Write header
        csv_writer.writerow(['index', 'sentence'])

        # Iterate over each row in the worksheet
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Extract sentence from the first column (column A in Excel)
            sentence = row[0]
            # Write index and sentence to CSV
            csv_writer.writerow([index, sentence])

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
print(sys.path[0])

# File paths
input_file = os.path.join(sys.path[0], "converters", "data", "german_sentences.xlsx")
output_file = os.path.join(sys.path[0], "de", "simple_sentences.csv" )

# Convert XLSX to CSV
xlsx_to_csv(input_file, output_file)

print("Converted from {} to {}".format(input_file, output_file))
